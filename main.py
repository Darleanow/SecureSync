from openpyxl import Workbook
from docx import Document
from re import compile, sub
from operator import itemgetter
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
from datetime import datetime

# reg to parse html
CLEANR = compile('<.*?>')

# Setting colors for the Excel
blue = "5abff2"
green = "92d050"
yellow = "f3d65b"
orange = "f99b4d"
red = "e95d5d"

# excel file creation
wb = Workbook()

# get active page ( first )
ws = wb.active

# setting tiles names
ws.title = "KERBEROS_report"
ws['A1'] = "Sujet"
ws['B1'] = "Catégorie"
ws['C1'] = "Présence Ping Castle"
ws['D1'] = "Présence Purple Knight"
ws['E1'] = "Criticité"
ws['F1'] = "Difficulté de rémédiation"
ws['G1'] = "Commentaire"
ws['H1'] = "Nom Flag Ping Castle"
ws['I1'] = "Score Ping Castle"
ws['J1'] = "Score Purple Knight"
ws['K1'] = "Astuce"
ws["L1"] = "Description Technique"
ws['M1'] = "Remédiation"
ws['N1'] = "Documentation"


# sets the PingCastle data to the Excel
def append_pc_data(ping_castle_data):
    # first we sort the data by score :
    cols = list(zip(*ping_castle_data))

    # Sorting by score
    cols.sort(key=itemgetter(7), reverse=True)

    # Zipping again
    sorted_pc_data = list(zip(*cols))

    # Iterating over it
    for i in range(len(sorted_pc_data[0])):
        number = i + 2

        # Subject
        ws.cell(row=number, column=1).value = sorted_pc_data[1][i]
        # Category
        ws.cell(row=number, column=2).value = sorted_pc_data[0][i]
        # Presence in PingCastle
        ws.cell(row=number, column=3).value = "Oui"
        # Presence in PurpleKnight
        ws.cell(row=number, column=4).value = "N/a"
        # Commentaire
        ws.cell(row=number, column=7).value = sorted_pc_data[4][i]
        # Flag name Ping Castle
        ws.cell(row=number, column=8).value = sorted_pc_data[3][i]
        # Score Ping Castle
        ws.cell(row=number, column=9).value = sorted_pc_data[7][i]
        # Score in PurpleKnight
        ws.cell(row=number, column=10).value = "N/a"
        # Additionnal information
        ws.cell(row=number, column=11).value = sorted_pc_data[2][i]
        # Technical description
        ws.cell(row=number, column=12).value = sorted_pc_data[5][i]
        # How to fix it
        ws.cell(row=number, column=13).value = sorted_pc_data[6][i]
        # Documentation
        ws.cell(row=number, column=14).value = "\n".join(i for i in sorted_pc_data[-1][i])

        # Setting color depending on the score
        if int(sorted_pc_data[7][i]) >= 15:
            for rows in ws.iter_rows(min_row=number, max_row=number, min_col=0, max_col=14):
                # Doing color for every cell in the rows
                for cell in rows:
                    cell.fill = PatternFill(start_color=red, end_color=red, fill_type="solid")
        elif int(sorted_pc_data[7][i]) >= 10:
            for rows in ws.iter_rows(min_row=number, max_row=number, min_col=0, max_col=14):
                for cell in rows:
                    cell.fill = PatternFill(start_color=orange, end_color=orange, fill_type="solid")
        elif int(sorted_pc_data[7][i]) >= 5:
            for rows in ws.iter_rows(min_row=number, max_row=number, min_col=0, max_col=14):
                for cell in rows:
                    cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")
        else:
            for rows in ws.iter_rows(min_row=number, max_row=number, min_col=0, max_col=14):
                for cell in rows:
                    cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
    # Saving file
    x = datetime.now()
    wb.save("excel_report_" + str(x.year) + "-" + str(x.month) + "-" + str(x.day) + ".xlsx")


# sets the PurpleKnight data to the Excel
def append_pk_data(purple_knight_data, append_start):
    # deziping list
    cols = list(zip(*purple_knight_data))

    # indexage
    cols.sort(key=itemgetter(5), reverse=True)

    # ziping again
    sorted_pk_data = list(zip(*cols))

    for i in range(len(sorted_pk_data[0])):
        # Starting at append_start because we don't want to overwrite the pingCastle data
        number = i + 2 + append_start

        """
        ws['A1'] = "Sujet"
        ws['B1'] = "Catégorie"
        ws['C1'] = "Présence Ping Castle"
        ws['D1'] = "Présence Purple Knight"
        ws['E1'] = "Criticité"
        ws['F1'] = "Difficulté de rémédiation"
        ws['G1'] = "Commentaire"
        ws['H1'] = "Nom Flag Ping Castle"
        ws['I1'] = "Score Ping Castle"
        ws['J1'] = "Score Purple Knight"
        ws['K1'] = "Astuce"
        ws["L1"] = "Description Technique"
        ws['M1'] = "Remédiation"
        ws['N1'] = "Documentation"
        """
        # [cat_name, sub_name, comment, tech_desc, remediation, weight]

        # subject
        ws.cell(row=number, column=1).value = sorted_pk_data[1][i]
        # Category
        ws.cell(row=number, column=2).value = sorted_pk_data[0][i]
        # PingCastle presence
        ws.cell(row=number, column=3).value = "N/a"
        # PurpleKnight presence
        ws.cell(row=number, column=4).value = "Oui"
        # Commentaire
        ws.cell(row=number, column=7).value = sorted_pk_data[2][i]
        # Flag name in PingCastle
        ws.cell(row=number, column=8).value = "N/a"
        # Score in PingCastle
        ws.cell(row=number, column=9).value = "N/a"
        # Score in Purple Knight
        ws.cell(row=number, column=10).value = str(sorted_pk_data[5][i])
        # Tip is none
        ws.cell(row=number,column=11).value = "None"
        # Tech desc
        ws.cell(row=number, column=12).value = sorted_pk_data[3][i]
        # Redemediation
        ws.cell(row=number, column=13).value = sorted_pk_data[4][i]
        # Documentation
        ws.cell(row=number, column=14).value = "None"

        # Setting colors depending on the score
        if int(sorted_pk_data[5][i]) >= 7:
            for rows in ws.iter_rows(min_row=number, max_row=number, min_col=0, max_col=14):
                for cell in rows:
                    cell.fill = PatternFill(start_color=red, end_color=red, fill_type="solid")
        elif int(sorted_pk_data[5][i]) >= 5:
            for rows in ws.iter_rows(min_row=number, max_row=number, min_col=0, max_col=14):
                for cell in rows:
                    cell.fill = PatternFill(start_color=orange, end_color=orange, fill_type="solid")
        elif int(sorted_pk_data[5][i]) >= 3:
            for rows in ws.iter_rows(min_row=number, max_row=number, min_col=0, max_col=14):
                for cell in rows:
                    cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")
        else:
            for rows in ws.iter_rows(min_row=number, max_row=number, min_col=0, max_col=14):
                for cell in rows:
                    cell.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
    # Saving Excel file
    x = datetime.now()
    wb.save("excel_report_" + str(x.year) + "-" + str(x.month) + "-" + str(x.day) + ".xlsx")


# Function to remove html tags
def cleanhtml(raw_html):
    cleantext = sub(CLEANR, '', raw_html)
    return cleantext


# Function to get the path of the reports
def get_files():
    ping_castle_path = input("Location complète du rapport PingCastle (HTML seulement): ")
    purple_knight_path = input("Location complète du rapport PurpleKnight (WORD seulement): ")
    return [ping_castle_path, purple_knight_path]


# Get the data of the PingCastle report
def extract_data_pc(ping_castle_report):
    # Initializing lists
    pc_category = list()
    pc_id_name = list()
    pc_tip = list()
    pc_rule_id = list()
    pc_desc = list()
    pc_tech = list()
    pc_solution = list()
    pc_score = list()
    pc_doc = list()

    # Default current category
    current_category = "Stale Objects"

    # Opening the file
    with open(ping_castle_report, "r", encoding='utf-8') as f:
        line = f.readline()
        # While there is data
        while line:
            if "<div" in line and "card-header" in line:
                line = f.readline()
                if "<h1" in line and "card-title" in line:
                    line = cleanhtml(line)
                    line = line.replace('\t', "").replace("\n", "")
                    current_category = line
            if "<span" in line and "card-title" in line:
                f.readline()
                line = f.readline()
                if "<" not in line and "}" not in line and len(line.split()) > 5 and "!=" not in line \
                        and "Objects" not in line:
                    line = line[0:-1]
                    line = line.lstrip()
                    pc_category.append(current_category)
                    pc_id_name.append(line)
            if "<div" in line and "card-body" in line:
                line = f.readline()
                if "section" not in line:
                    newline = cleanhtml(line)
                    newline = newline.lstrip()
                    newline = newline[0:-1]
                    if len(newline) > 11:
                        pc_tip.append(newline)
            if "<strong" in line and "Rule" in line:
                line = cleanhtml(line).replace("\n", "").replace("Rule ID:", "")
                pc_rule_id.append(line)
            if "<strong" in line and "Description" in line:
                line = cleanhtml(line).replace("\n", "").replace("Description:", "")
                pc_desc.append(line)
            if "<strong" in line and "Technical" in line:
                line = cleanhtml(line).replace("\n", "").replace("Technical explanation:", "")
                pc_tech.append(line)
            if "<strong" in line and "Advised" in line:
                line = cleanhtml(line).replace("\n", "").replace("Advised solution:", "")
                pc_solution.append(line)
            if "<strong" in line and "Points" in line:
                line = cleanhtml(line).replace("\n", "").replace("Points:", "")
                if "Informative" in line:
                    line = "0"
                else:
                    line = line.split()
                    line = line[0]
                if line == "5":
                    line = "05"
                if line == "1":
                    line = "01"
                pc_score.append(line)
            if "<strong" in line and "Documentation" in line:
                tmp_lst = list()
                while "<a href=" in line:
                    soup = BeautifulSoup(line, "html.parser").find_all(lambda t: t.name == "a")
                    tmp_tmp_lst = [a["href"] for a in soup if len(a["href"]) > 10]
                    tmp_lst.append(tmp_tmp_lst[0])
                    line = f.readline()
                pc_doc.append(tmp_lst)
            line = f.readline()
    return [pc_category, pc_id_name, pc_tip, pc_rule_id, pc_desc, pc_tech, pc_solution, pc_score, pc_doc]


# get the data of the PurpleKnight report
def extract_data_pk(pk_report_url_data):
    # Open the Word document
    document = Document(pk_report_url_data)

    # Initializing lists
    sub_name = list()
    cat_name = list()
    weight = list()
    comment = list()
    tech_desc = list()
    remediation = list()

    # As there is no html in this, set bools for when it's needed to change values
    change_subname = False
    change_category = False
    change_weight = False
    change_comment = False
    change_tech_desc = False
    change_remediation = False

    # Initializing string
    current_subject = ""
    current_category = "Account Security"
    current_weight = ""
    current_comment = ""
    current_tech_desc = ""
    current_remediation = ""

    # Iterate through each paragraph in the document
    for paragraph in document.paragraphs:
        # Print the text of the paragraph
        if change_category == True and len(str(paragraph.text)) != 0:
            current_category = str(paragraph.text).lower().capitalize()
            change_category = False
        if "CATEGORY" in str(paragraph.text):
            change_category = True
        if change_weight:
            current_weight = str(paragraph.text)
            change_weight = False
        if "WEIGHT" in str(paragraph.text):
            change_weight = True
        if change_subname:
            tmp_str = str(paragraph.text)
            tmp_str = tmp_str.split("IOE Found")
            current_subject = tmp_str[0]
            change_subname = False
        if "SECURITY INDICATOR" in str(paragraph.text):
            change_subname = True
        if change_comment:
            current_comment = str(paragraph.text)
            change_comment = False
        if "Description" in str(paragraph.text):
            change_comment = True
        if change_tech_desc:
            current_tech_desc = str(paragraph.text)
            change_tech_desc = False
        if "Likelihood of Compromise" in str(paragraph.text):
            change_tech_desc = True
        if change_remediation:
            current_remediation = str(paragraph.text)
            change_remediation = False
        if "Remediation Steps" in str(paragraph.text):
            change_remediation = True
        # if every value is set, append the line
        if len(current_subject) > 0 and len(current_remediation) > 0 and len(current_comment) > 0 and \
                len(current_weight) > 0 and len(current_tech_desc) > 0:
            sub_name.append(current_subject)
            cat_name.append(current_category)
            weight.append(int(current_weight))
            comment.append(current_comment)
            tech_desc.append(current_tech_desc)
            remediation.append(current_remediation)

            current_remediation, current_subject, current_comment, current_weight, current_tech_desc = \
                "", "", "", "", ""
    return [cat_name, sub_name, comment, tech_desc, remediation, weight]


if __name__ == "__main__":
    # get paths
    files = get_files()
    # setting variable names
    pc_report_url, pk_report_url = files[0], files[1]
    # Informational print
    print("!! Scraping data from file. . .")
    # Extract data
    pc_data = extract_data_pc(pc_report_url)
    pk_data = extract_data_pk(pk_report_url)
    # Sucess if no tracebacks
    print("> Sucess scraping data from file. . .")
    # informational print
    print("!! Adding data to excel. . .")
    # Appending data from pc and pk
    append_pc_data(pc_data)
    append_pk_data(pk_data, len(pc_data[0]))
    # Success !
    print("> Success adding data to excel, check the file, correct if needed, and then press enter. . .")
    input("")
